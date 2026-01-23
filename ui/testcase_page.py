
import openpyxl
from app_state import app_context
from excel_exporter import export_service_recommendations
from service_aggregation import aggregate_service_rows
from PySide6.QtWidgets import (
    QWidget,
    QTableWidget,
    QTableWidgetItem,
    QPushButton,
    QLabel,
    QFrame,
    QHBoxLayout,
    QVBoxLayout,
    QSizePolicy,
    QHeaderView,
    QAbstractItemView,
    QFileDialog,
    QMessageBox,
)
from PySide6.QtUiTools import QUiLoader
from PySide6.QtCore import QFile, Qt

from utils.path_helper import resource_path

class TestCasePage(QWidget):
    def __init__(self):
        super().__init__()

        loader = QUiLoader()
        file = QFile(str(resource_path("ui/testcase_page.ui")))
        file.open(QFile.ReadOnly)
        self.ui = loader.load(file, self)
        file.close()

        self.table = self.ui.findChild(QTableWidget, "tableTestCases")
        self.btn_import = self.ui.findChild(QPushButton, "btnImport")
        self.btn_export = self.ui.findChild(QPushButton, "btnExport")
        self.service_groups = []

        self._style_table()
        model = self.table.model()
        model.rowsInserted.connect(lambda *_: self.apply_row_styling())
        model.modelReset.connect(self.apply_row_styling)
        model.layoutChanged.connect(self.apply_row_styling)
        self.table.selectionModel().selectionChanged.connect(
            lambda *_: self._sync_widget_selection()
        )

        self.load_mock_data()
        self.apply_row_styling()

        if self.btn_import:
            self.btn_import.clicked.connect(self.import_excel)
        if self.btn_export:
            self.btn_export.clicked.connect(self.export_to_excel)

    def load_mock_data(self):
        sample_rows = [
            {
                "service": "Auth Gateway",
                "priority": "High",
                "security": "High",
                "country": "All",
                "mobile": True,
                "web_console": True,
                "actions": [],
                "tenant_requirement": "Common",
            },
            {
                "service": "Payments",
                "priority": "Medium",
                "security": "Medium",
                "country": "All except HKO",
                "mobile": False,
                "web_console": True,
                "actions": [],
                "tenant_requirement": "Common HKO",
            },
            {
                "service": "Notifications",
                "priority": "Low",
                "security": "Low",
                "country": "HKO",
                "mobile": False,
                "web_console": False,
                "actions": [],
                "tenant_requirement": "HKO",
            },
        ]

        rows, groups = aggregate_service_rows(sample_rows)

        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, value in enumerate(row):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(r, c, item)

        self.service_groups = groups
        app_context["service_groups"] = groups
        app_context["raw_rows"] = sample_rows
        app_context["ai_results"] = []
        app_context["service_groups"] = self.service_groups
        app_context["ai_results"] = []

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Upload Excel File",
            "",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            self._import_excel_from_path(path)
        except Exception as exc:
            QMessageBox.critical(self, "Import Failed", str(exc))

    def _import_excel_from_path(self, path: str):
        processed = self._load_excel_file(path)
        if not processed:
            return

        rows, groups, raw_rows, header_names = processed
        if not rows:
            return

        self.service_groups = groups
        app_context["service_groups"] = groups
        app_context["raw_rows"] = raw_rows
        app_context["raw_headers"] = header_names
        app_context["ai_results"] = []
        self.table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_index, col_index, item)

        self.apply_row_styling()
        self.table.resizeColumnsToContents()

    def export_to_excel(self):
        groups = app_context.get("service_groups", [])
        ai_results = app_context.get("ai_results", [])

        if not groups:
            QMessageBox.information(
                self,
                "Export Test Case",
                "Load test case data before exporting.",
            )
            return

        if not ai_results:
            QMessageBox.information(
                self,
                "Export Test Case",
                "Please generate AI recommendation before exporting.",
            )
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Test Case Export",
            "testcase_export.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return

        try:
            export_service_recommendations(groups, ai_results, path)
            QMessageBox.information(self, "Export Test Case", "Exported successfully.")
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))

    def _load_excel_file(self, path: str):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return [], [], []

        sheet = wb[wb.sheetnames[0]]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if not header_row:
            return [], [], []

        header_names = [str(val).strip() for val in header_row if val is not None]
        headers = {
            str(val).strip().lower(): idx
            for idx, val in enumerate(header_row)
            if val is not None
        }

        required_columns = {
            "service",
            "ash test scenarios",
            "priority",
            "security priority",
            "evidence mobile",
            "evidence web console",
            "countries",
            "action",
            "tenant requirement",
        }
        if not required_columns.issubset(headers.keys()):
            return [], [], []

        raw_rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            service = (
                str(row[headers["service"]]).strip()
                if row[headers["service"]] is not None
                else ""
            )
            if not service:
                continue

            action = (
                str(row[headers["action"]]).strip()
                if row[headers["action"]] is not None
                else ""
            )
            priority = (
                str(row[headers["priority"]]).strip()
                if row[headers["priority"]] is not None
                else "Medium"
            )
            security = (
                str(row[headers["security priority"]]).strip()
                if row[headers["security priority"]] is not None
                else "Medium"
            )
            mobile = (
                str(row[headers["evidence mobile"]]).strip().upper()
                if row[headers["evidence mobile"]] is not None
                else "N"
            )
            web_console = (
                str(row[headers["evidence web console"]]).strip().upper()
                if row[headers["evidence web console"]] is not None
                else "N"
            )
            country = (
                str(row[headers["countries"]]).strip()
                if row[headers["countries"]] is not None
                else ""
            )
            tenant_requirement = (
                str(row[headers["tenant requirement"]]).strip()
                if row[headers["tenant requirement"]] is not None
                else ""
            )
            need_actual_car = (
                str(row[headers["need actual car"]]).strip().upper()
                if row[headers["need actual car"]] is not None
                else "N"
            )

            raw_rows.append(
                {
                    "service": service,
                    "priority": priority,
                    "security": security,
                    "country": country,
                    "mobile": mobile == "Y",
                    "web_console": web_console == "Y",
                    "actions": [action] if action else [],
                    "tenant_requirement": tenant_requirement,
                    "need_actual_car": need_actual_car == "Y",
                }
            )

        rows, groups = aggregate_service_rows(raw_rows)
        return rows, groups, raw_rows, header_names

    def _style_table(self):
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)
        self.table.setFrameShape(QFrame.NoFrame)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.verticalHeader().setVisible(False)

        header = self.table.horizontalHeader()
        header.setHighlightSections(False)
        header.setDefaultAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        SERVICE_COL = 1
        PRIORITY_COL = 3
        SECURITY_COL = 4
        COUNTRY_COL = 7

        header.setSectionResizeMode(SERVICE_COL, QHeaderView.Stretch)
        header.setSectionResizeMode(COUNTRY_COL, QHeaderView.Stretch)
        # DPI-safe: keep Priority/Security columns consistent width
        header.setSectionResizeMode(PRIORITY_COL, QHeaderView.Fixed)
        header.setSectionResizeMode(SECURITY_COL, QHeaderView.Fixed)
        self.table.setColumnWidth(PRIORITY_COL, 140)
        self.table.setColumnWidth(SECURITY_COL, 140)
        # DPI-safe: allow header to grow if needed
        header.setMinimumHeight(44)
        header.setStyleSheet(
            "QHeaderView::section {"
            "background-color: #3f3f3f;"
            "color: white;"
            "font-weight: 600;"
            "border: none;"
            "padding: 12px 12px;"
            "}"
        )

        self.table.setStyleSheet(
            """
            QTableWidget {
                background-color: white;
                border-radius: 14px;
                border: none;
                alternate-background-color: #fbfbfb;
            }
            QTableWidget::item {
                border-bottom: 1px solid #f0f0f0;
                padding: 0;
            }
            QTableWidget::item:selected {
                background-color: #0b74d1;
                color: white;
            }
            QTableWidget::item:selected:!active {
                background-color: #0b74d1;
            }
            """
        )

    def apply_row_styling(self):
        for row in range(self.table.rowCount()):
            self.table.setRowHeight(row, 54)
            self._update_service_badge(row)
            self._update_priority_cell(row)
            self._update_security_cell(row)
            self._update_boolean_cell(row, 5)
            self._update_boolean_cell(row, 6)
        # setCellWidget does not inherit selection styles automatically
        self._sync_widget_selection()

    def _update_service_badge(self, row: int):
        service_col = 1
        if service_col >= self.table.columnCount():
            return
        item = self.table.item(row, service_col)
        if not item:
            return
        text = item.text().lower().strip()
        if text not in ("manual", "auto"):
            return
        badge = self._create_badge(text)
        self.table.setCellWidget(row, service_col, badge)

    def _create_badge(self, text: str) -> QLabel:
        colors = {"manual": "#d32f2f", "auto": "#9e9e9e"}
        label = QLabel(text.capitalize())
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet(
            f"""
            background-color: {colors.get(text, '#9e9e9e')};
            color: white;
            font-weight: 600;
            border-radius: 14px;
            padding: 3px 16px;
            """
        )
        label.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
        return label

    def _update_priority_cell(self, row: int):
        priority_col = 3
        if priority_col >= self.table.columnCount():
            return
        item = self.table.item(row, priority_col)
        if not item:
            return
        widget = self._create_progress_widget(item.text(), show_label=False)
        self.table.setCellWidget(row, priority_col, self._center_cell(widget))
        item.setText("")

    def _update_security_cell(self, row: int):
        security_col = 4
        item = self.table.item(row, security_col)
        if not item:
            return

        widget = self._create_progress_widget(item.text(), show_label=False)
        self.table.setCellWidget(row, security_col, self._center_cell(widget))
        item.setText("")

    def _update_boolean_cell(self, row: int, column: int):
        item = self.table.item(row, column)
        if not item:
            return

        widget = self._create_boolean_indicator(
            item.text().strip().upper() == "Y"
        )
        self.table.setCellWidget(row, column, widget)
        item.setText("")

    def _create_progress_widget(self, level_text: str, show_label: bool) -> QWidget:
        level = level_text.strip().lower()

        mapping = {
            "high": (95, "#d32f2f"),
            "medium": (65, "#ff9800"),
            "low": (35, "#fbc02d"),
        }

        percent, color = mapping.get(level, (45, "#90a4ae"))

        # --- Track ---
        track = QFrame()
        track.setMinimumHeight(10)
        track.setStyleSheet(
            "background-color: #e1e1e1; border-radius: 5px;"
        )

        track_layout = QHBoxLayout(track)
        track_layout.setContentsMargins(0, 0, 0, 0)
        track_layout.setSpacing(0)

        # --- Indicator ---
        indicator = QFrame()
        indicator.setStyleSheet(
            f"""
            background-color: {color};
            border-radius: 5px;
            """
        )

        # ⭐ key fix: ใช้ stretch แทน fixed width
        track_layout.addWidget(indicator, percent)
        track_layout.addStretch(100 - percent)

        # --- Center container (for cell alignment) ---
        center = QWidget()
        center.setStyleSheet("background: transparent;")
        center_layout = QHBoxLayout(center)
        center_layout.setContentsMargins(0, 0, 0, 0)
        center_layout.addStretch()
        center_layout.addWidget(track)
        center_layout.addStretch()
       
        if show_label:
            wrapper = QWidget()
            wrapper.setStyleSheet("background: transparent;")
            layout = QVBoxLayout(wrapper)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(6)

            label = QLabel(level_text.capitalize())
            label.setStyleSheet("color: #2a2a2a; font-weight: 600;")
            label.setAlignment(Qt.AlignCenter)

            layout.addWidget(label)
            layout.addWidget(center)
            return wrapper

        return center

    def _center_cell(self, widget: QWidget) -> QWidget:
        """Center custom widgets to prevent DPI drift in fixed-width columns."""
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignCenter)
        layout.addWidget(widget)
        return container

    def _sync_widget_selection(self):
        selected_rows = {idx.row() for idx in self.table.selectionModel().selectedRows()}
        selected_color = "#0b74d1"
        for row in range(self.table.rowCount()):
            is_selected = row in selected_rows
            for col in (3, 4):
                widget = self.table.cellWidget(row, col)
                if widget is None:
                    continue
                if is_selected:
                    widget.setStyleSheet(f"background-color: {selected_color};")
                else:
                    widget.setStyleSheet("background-color: transparent;")



    def _create_boolean_indicator(self, checked: bool) -> QWidget:
        label = QLabel("✓" if checked else "✕")
        label.setAlignment(Qt.AlignCenter)
        label.setMinimumSize(20, 20)
        label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        if checked:
            label.setStyleSheet("""
                color: white;
                background-color: #4caf50;
                border-radius: 10px;
                font-weight: bold;
            """)
        else:
            label.setStyleSheet("""
                color: white;
                background-color: #bdbdbd;
                border-radius: 10px;
                font-weight: bold;
            """)

        container = QWidget()
        container.setStyleSheet("background: transparent;")
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addStretch()
        layout.addWidget(label)
        layout.addStretch()
        return container
