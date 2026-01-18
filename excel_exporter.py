import openpyxl
from openpyxl.styles import Alignment


def _build_ai_text(ai_result: dict) -> str:
    if ai_result.get("forced_manual"):
        return f"Risk {ai_result['risk_score']}% | Manual 100% (Forced)"
    return (
        f"Risk {ai_result['risk_score']}% | Manual {ai_result['manual_percent']}% / "
        f"Auto {ai_result['auto_percent']}%"
    )


def export_service_recommendations(groups: list, ai_results: list, path: str) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    headers = [
        "No",
        "Service",
        "Number of Scenario",
        "Priority",
        "Security",
        "Country",
        "Mobile",
        "Web Console",
        "AI Recommend to Test",
    ]
    worksheet.append(headers)

    for idx, (group, ai_result) in enumerate(zip(groups, ai_results), start=1):
        ai_text = _build_ai_text(ai_result)
        row = [
            str(idx),
            group.get("service", ""),
            str(group.get("scenarios", "")),
            group.get("priority", ""),
            group.get("security", ""),
            group.get("country", ""),
            "Y" if group.get("mobile") else "N",
            "Y" if group.get("web_console") else "N",
            ai_text,
        ]
        worksheet.append(row)
        current_row = worksheet.max_row
        for col_index in range(1, len(row) + 1):
            worksheet.cell(row=current_row, column=col_index).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    workbook.save(path)


def export_ai_recommend_excel(
    rows: list,
    risk_scores: list,
    ratio_summary: str,
    output_path: str,
) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    headers = [
        "No",
        "Service",
        "ASH Test Scenarios",
        "Priority",
        "Security Priority",
        "Evidence Mobile",
        "Evidence Web Console",
        "Countries",
        "Need Actual car",
        "Action",
        "Tenant Requirement",
        "AI Risk Score (%)",
        "AI Recommend Test Type",
        "AI Recommend Ratio Summary",
    ]
    worksheet.append(headers)

    manual_actions = {"test drive", "vehicle support", "application"}

    for index, row in enumerate(rows, start=1):
        actions = row.get("actions") or []
        if isinstance(actions, str):
            actions_list = [actions]
        else:
            actions_list = list(actions)
        action_text = "; ".join(actions_list)
        test_type = (
            "Manual"
            if any(action.strip().lower() in manual_actions for action in actions_list)
            else "Automation"
        )
        risk_score = risk_scores[index - 1] if index - 1 < len(risk_scores) else 0

        data_row = [
            str(index),
            row.get("service", ""),
            str(row.get("scenarios", "")),
            row.get("priority", ""),
            row.get("security", ""),
            "Y" if row.get("mobile") else "N",
            "Y" if row.get("web_console") else "N",
            row.get("country", ""),
            "Y" if row.get("need_actual_car") else "N",
            action_text,
            row.get("tenant_requirement", ""),
            risk_score,
            test_type,
            ratio_summary,
        ]
        worksheet.append(data_row)
        current_row = worksheet.max_row
        for col_index in range(1, len(data_row) + 1):
            worksheet.cell(row=current_row, column=col_index).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    workbook.save(output_path)
